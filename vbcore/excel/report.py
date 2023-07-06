import traceback as tbk
from collections import Counter
from dataclasses import dataclass
from enum import auto, Enum
from typing import Any, ClassVar, Generic, List, Optional, Type, TypeVar

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

from vbcore.base import BaseDTO
from vbcore.enums import StrEnum
from vbcore.types import OptExc, StrDict, TripleStr

R = TypeVar("R", bound="ReportRow")


@dataclass(frozen=True)
class ReportSettings:
    error_tab_color: str = "DD0000"
    warning_tab_color: str = "FF9F00"
    info_tab_color: str = "00DD00"
    header_border: Border = Border()
    header_alignment: Alignment = Alignment(horizontal="center", vertical="center")
    header_font: Font = Font(color="000000", bold=True)
    header_fill: PatternFill = PatternFill(
        start_color=Color(indexed=22),
        end_color=Color(indexed=22),
        fill_type="solid",
    )
    cell_border: Border = Border(
        top=Side(border_style="thin", color="454545"),
        left=Side(border_style="thin", color="454545"),
        right=Side(border_style="thin", color="454545"),
        bottom=Side(border_style="thin", color="454545"),
    )


class Level(StrEnum):
    INFO = auto()
    WARN = auto()
    ERROR = auto()


@dataclass(frozen=True)
class Event:
    level: Level
    data: dict


@dataclass(frozen=True)
class ReportRow(BaseDTO):
    ALIASES: ClassVar[StrDict] = {}

    @classmethod
    def alias(cls, field: str) -> str:
        return cls.ALIASES.get(field) or field

    @classmethod
    def decode_exception(cls, exc: Exception) -> TripleStr:
        message = str(exc)
        name = exc.__class__.__name__
        stacktrace = tbk.format_exception(None, exc, exc.__traceback__)
        return name, message, "".join(stacktrace).strip()

    def from_exception(self: R, exc: Exception) -> R:
        field, message, stacktrace = self.decode_exception(exc)
        return self(
            field=field,
            stacktrace=stacktrace,
            message=getattr(self, "message", None) or message,
        )


class Report(Generic[R]):
    def __init__(
        self,
        row_class: Optional[Type[R]] = None,
        settings: Optional[ReportSettings] = None,
        max_rows: int = 2**18,
        max_col_width: int = 999,
    ):
        self.max_rows = max_rows
        self.max_col_width = max_col_width
        self.row_class = row_class or ReportRow
        self._events: List[Event] = []
        self.settings = settings or ReportSettings()
        self._workbook: Workbook = Workbook()
        self._error_sheet = self._workbook.active
        self._error_sheet.title = Level.ERROR.value
        self._warning_sheet = self._workbook.create_sheet(Level.WARN.value, 1)
        self._info_sheet = self._workbook.create_sheet(Level.INFO.value, 2)
        self._error_sheet.sheet_properties.tabColor = self.settings.error_tab_color
        self._warning_sheet.sheet_properties.tabColor = self.settings.warning_tab_color
        self._info_sheet.sheet_properties.tabColor = self.settings.info_tab_color

    def get_sheet(self, level: Level):
        sheet_maps = {
            Level.INFO: self._info_sheet,
            Level.WARN: self._warning_sheet,
            Level.ERROR: self._error_sheet,
        }
        return sheet_maps[level]

    def register_event(
        self, level: Level, *, row: Optional[R] = None, exc: OptExc = None, **kwargs
    ):
        record = row or self.row_class()
        record = record(**kwargs) if exc is None else record.from_exception(exc)

        if len(self._events) < self.max_rows:
            self._events.append(Event(level, record.to_dict()))

    def error(self, row: Optional[R] = None, exc: OptExc = None, **kwargs):
        self.register_event(Level.ERROR, row=row, exc=exc, **kwargs)

    def warning(self, row: Optional[R] = None, exc: OptExc = None, **kwargs):
        self.register_event(Level.WARN, row=row, exc=exc, **kwargs)

    def info(self, row: Optional[R] = None, exc: OptExc = None, **kwargs):
        self.register_event(Level.INFO, row=row, exc=exc, **kwargs)

    def _adjust_columns_width(self):
        for level in Level:
            dimensions = {}
            sheet = self.get_sheet(level)
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        dimensions[cell.column_letter] = min(
                            self.max_col_width,
                            max(
                                dimensions.get(cell.column_letter, 0),
                                len(str(cell.value)),
                            ),
                        )
            for col, value in dimensions.items():
                sheet.column_dimensions[col].width = value + 3

    def prepare_header_cells(self):
        for index, header in enumerate(self.row_class.field_names(), 1):
            column_header = self.row_class.alias(header)
            for level in Level:
                sheet = self.get_sheet(level)
                header_cell = sheet.cell(1, index, column_header)
                header_cell.font = self.settings.header_font
                header_cell.border = self.settings.header_border
                header_cell.fill = self.settings.header_fill
                header_cell.alignment = self.settings.header_alignment
                header_cell.border = self.settings.cell_border

    @classmethod
    def decode_value(cls, value: Any) -> str:
        """
        Use this method to format different value types
        until a better way is found
        """
        if isinstance(value, Enum):
            return str(value.value)
        return "" if value is None else str(value)

    def prepare_row_cells(self):
        row_counter = Counter({level: 2 for level in Level})
        for event in self._events:
            sheet = self.get_sheet(event.level)
            for index, header in enumerate(self.row_class.field_names(), 1):
                value = self.decode_value(event.data.get(header, ""))
                cell = sheet.cell(row_counter[event.level], index, value)
                cell.border = self.settings.cell_border
            row_counter[event.level] += 1

    def write_report(self, filename: str):
        if len(self._events) >= self.max_rows - 1:
            self.error(
                message="CRITICAL: Reached the maximum number of lines. "
                "There may be unreported information. Check the logs!",
            )

        self.prepare_header_cells()
        self.prepare_row_cells()
        self._adjust_columns_width()
        self._workbook.save(filename)
