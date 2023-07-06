from dataclasses import dataclass

from openpyxl import load_workbook

from vbcore.excel.report import Level, Report, ReportRow
from vbcore.types import OptStr


@dataclass(frozen=True)
class CustomReportRow(ReportRow):
    ALIASES = {
        "test1": "col1",
        "test2": "col2",
        "test3": "col3",
    }

    test1: OptStr = None
    test2: OptStr = None
    test3: OptStr = None
    field: OptStr = None
    message: OptStr = None


def test_excel_report(tmpdir):
    file = tmpdir.join("test_excel_report.xlsx")
    report = Report[CustomReportRow](row_class=CustomReportRow)

    report.info(test1="AAA", test2="BBB")
    report.error(test2="CCC", test3="DDD")
    report.warning(test1="EEE", test4="FFF")
    report.error(exc=ValueError("value error"))
    report.write_report(file.strpath)

    wbk = load_workbook(file.strpath)
    assert wbk.sheetnames == [Level.ERROR.value, Level.WARN.value, Level.INFO.value]
    assert wbk.active.title == Level.ERROR.value

    sheet_error = wbk[Level.ERROR.value]
    assert sheet_error.cell(1, 1).value == "col1"
    assert sheet_error.cell(1, 2).value == "col2"
    assert sheet_error.cell(1, 3).value == "col3"
    assert sheet_error.cell(1, 4).value == "field"
    assert sheet_error.cell(1, 5).value == "message"
    assert sheet_error.cell(2, 1).value is None
    assert sheet_error.cell(2, 2).value == "CCC"
    assert sheet_error.cell(2, 3).value == "DDD"
    assert sheet_error.cell(3, 1).value is None
    assert sheet_error.cell(3, 2).value is None
    assert sheet_error.cell(3, 3).value is None
    assert sheet_error.cell(3, 4).value == "ValueError"
    assert sheet_error.cell(3, 5).value == "value error"

    sheet_warn = wbk[Level.WARN.value]
    assert sheet_warn.cell(1, 1).value == "col1"
    assert sheet_warn.cell(1, 2).value == "col2"
    assert sheet_warn.cell(1, 3).value == "col3"
    assert sheet_warn.cell(1, 4).value == "field"
    assert sheet_warn.cell(1, 5).value == "message"
    assert sheet_warn.cell(2, 1).value == "EEE"
    assert sheet_warn.cell(2, 2).value is None
    assert sheet_warn.cell(2, 3).value is None

    sheet_info = wbk[Level.INFO.value]
    assert sheet_info.cell(1, 1).value == "col1"
    assert sheet_info.cell(1, 2).value == "col2"
    assert sheet_info.cell(1, 3).value == "col3"
    assert sheet_info.cell(1, 4).value == "field"
    assert sheet_info.cell(1, 5).value == "message"
    assert sheet_info.cell(2, 1).value == "AAA"
    assert sheet_info.cell(2, 2).value == "BBB"
    assert sheet_info.cell(2, 3).value is None
